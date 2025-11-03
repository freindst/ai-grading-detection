"""
Export Manager - Handle export of grading results in various formats
"""

import csv
import json
from typing import List, Dict, Optional
from pathlib import Path
from datetime import datetime


class ExportManager:
    """Manage export of grading results to different formats"""
    
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
    
    def export_to_csv(
        self,
        results: List[Dict],
        output_file: str,
        include_full_feedback: bool = True
    ) -> tuple[bool, str]:
        """
        Export results to CSV file
        
        Args:
            results: List of grading results
            output_file: Output file path
            include_full_feedback: Whether to include full feedback text
            
        Returns:
            Tuple of (success, message)
        """
        try:
            output_path = self.export_dir / output_file
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                if include_full_feedback:
                    fieldnames = [
                        'Filename', 'Grade', 'Confidence',
                        'Detailed Feedback', 'Student Feedback',
                        'Strengths', 'Weaknesses',
                        'AI Keywords Found', 'Plagiarism Status',
                        'Status', 'Error'
                    ]
                else:
                    fieldnames = [
                        'Filename', 'Grade', 'Confidence',
                        'Strengths Count', 'Weaknesses Count',
                        'AI Keywords Found', 'Plagiarism Status',
                        'Status'
                    ]
                
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in results:
                    # Determine plagiarism status
                    plag_status = "None"
                    if result.get('plagiarism_pairs'):
                        max_sim = max((p['similarity'] for p in result['plagiarism_pairs']), default=0)
                        if max_sim >= 80:
                            plag_status = f"High ({max_sim}%)"
                        elif max_sim >= 60:
                            plag_status = f"Medium ({max_sim}%)"
                        else:
                            plag_status = f"Low ({max_sim}%)"
                    
                    if include_full_feedback:
                        row = {
                            'Filename': result.get('filename', ''),
                            'Grade': result.get('grade', 'N/A'),
                            'Confidence': result.get('confidence', 'N/A'),
                            'Detailed Feedback': result.get('detailed_feedback', '')[:1000],
                            'Student Feedback': result.get('student_feedback', '')[:1000],
                            'Strengths': '; '.join(result.get('strengths', [])),
                            'Weaknesses': '; '.join(result.get('weaknesses', [])),
                            'AI Keywords Found': '; '.join(result.get('ai_keywords_found', [])),
                            'Plagiarism Status': plag_status,
                            'Status': 'Success' if result.get('success') else 'Failed',
                            'Error': result.get('error', '')
                        }
                    else:
                        row = {
                            'Filename': result.get('filename', ''),
                            'Grade': result.get('grade', 'N/A'),
                            'Confidence': result.get('confidence', 'N/A'),
                            'Strengths Count': len(result.get('strengths', [])),
                            'Weaknesses Count': len(result.get('weaknesses', [])),
                            'AI Keywords Found': '; '.join(result.get('ai_keywords_found', [])),
                            'Plagiarism Status': plag_status,
                            'Status': 'Success' if result.get('success') else 'Failed'
                        }
                    
                    writer.writerow(row)
            
            return True, f"Exported to {output_path}"
        
        except Exception as e:
            return False, f"CSV export failed: {str(e)}"
    
    def export_to_json(
        self,
        results: List[Dict],
        output_file: str,
        pretty: bool = True
    ) -> tuple[bool, str]:
        """Export results to JSON file"""
        try:
            output_path = self.export_dir / output_file
            
            # Simplify results for export
            export_data = []
            for result in results:
                simplified = {
                    "filename": result.get('filename'),
                    "grade": result.get('grade'),
                    "confidence": result.get('confidence'),
                    "detailed_feedback": result.get('detailed_feedback'),
                    "student_feedback": result.get('student_feedback'),
                    "strengths": result.get('strengths'),
                    "weaknesses": result.get('weaknesses'),
                    "ai_keywords_found": result.get('ai_keywords_found'),
                    "plagiarism_pairs": result.get('plagiarism_pairs'),
                    "success": result.get('success'),
                    "error": result.get('error')
                }
                export_data.append(simplified)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                if pretty:
                    json.dump(export_data, f, indent=2, ensure_ascii=False)
                else:
                    json.dump(export_data, f, ensure_ascii=False)
            
            return True, f"Exported to {output_path}"
        
        except Exception as e:
            return False, f"JSON export failed: {str(e)}"
    
    def export_to_xlsx(
        self,
        results: List[Dict],
        output_file: str
    ) -> tuple[bool, str]:
        """Export results to Excel file"""
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                return False, "openpyxl not installed. Run: pip install openpyxl"
            
            output_path = self.export_dir / output_file
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Grading Results"
            
            # Header
            headers = ['Filename', 'Grade', 'Confidence', 'Detailed Feedback', 
                      'Student Feedback', 'Strengths', 'Weaknesses', 
                      'AI Keywords', 'Plagiarism', 'Status']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data rows
            for row_idx, result in enumerate(results, 2):
                plag_status = "None"
                if result.get('plagiarism_pairs'):
                    max_sim = max((p['similarity'] for p in result['plagiarism_pairs']), default=0)
                    plag_status = f"{max_sim}%"
                
                row_data = [
                    result.get('filename', ''),
                    result.get('grade', 'N/A'),
                    result.get('confidence', 'N/A'),
                    result.get('detailed_feedback', ''),
                    result.get('student_feedback', ''),
                    '\n'.join(result.get('strengths', [])),
                    '\n'.join(result.get('weaknesses', [])),
                    ', '.join(result.get('ai_keywords_found', [])),
                    plag_status,
                    'Success' if result.get('success') else 'Failed'
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_path)
            return True, f"Exported to {output_path}"
        
        except Exception as e:
            return False, f"Excel export failed: {str(e)}"
    
    def export_summary_stats(
        self,
        results: List[Dict],
        output_file: str
    ) -> tuple[bool, str]:
        """Export summary statistics as text file"""
        try:
            output_path = self.export_dir / output_file
            
            # Calculate statistics
            total = len(results)
            successful = sum(1 for r in results if r.get('success'))
            failed = total - successful
            
            # Grade distribution
            grade_dist = {}
            for result in results:
                if result.get('success'):
                    grade = result.get('grade', 'N/A')
                    grade_dist[grade] = grade_dist.get(grade, 0) + 1
            
            # Plagiarism summary
            high_plag = 0
            medium_plag = 0
            for result in results:
                if result.get('plagiarism_pairs'):
                    max_sim = max((p['similarity'] for p in result['plagiarism_pairs']), default=0)
                    if max_sim >= 80:
                        high_plag += 1
                    elif max_sim >= 60:
                        medium_plag += 1
            
            # Write summary
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 60 + "\n")
                f.write("GRADING SUMMARY REPORT\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 60 + "\n\n")
                
                f.write(f"Total Submissions: {total}\n")
                f.write(f"Successfully Graded: {successful}\n")
                f.write(f"Failed: {failed}\n")
                f.write(f"Success Rate: {(successful/total*100):.1f}%\n\n")
                
                f.write("GRADE DISTRIBUTION:\n")
                f.write("-" * 30 + "\n")
                for grade, count in sorted(grade_dist.items()):
                    percentage = (count / successful * 100) if successful > 0 else 0
                    f.write(f"  {grade:5s}: {count:3d} ({percentage:5.1f}%)\n")
                
                f.write("\nPLAGIARISM SUMMARY:\n")
                f.write("-" * 30 + "\n")
                f.write(f"  High Suspicion (≥80%): {high_plag}\n")
                f.write(f"  Medium Suspicion (≥60%): {medium_plag}\n")
                
                f.write("\n" + "=" * 60 + "\n")
            
            return True, f"Summary exported to {output_path}"
        
        except Exception as e:
            return False, f"Summary export failed: {str(e)}"
    
    def generate_filename(self, prefix: str, extension: str) -> str:
        """Generate timestamped filename"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"

